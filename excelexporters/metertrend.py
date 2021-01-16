import base64
import uuid
import os
from openpyxl.chart import (
    PieChart,
    BarChart,
    Reference,
)
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList


####################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file bytes to Base64
####################################################################################################################


def export(result,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    if "reporting_period" not in result.keys() or \
            "names" not in result['reporting_period'].keys() or len(result['reporting_period']['names']) == 0:
        return None
    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type)

    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(result,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type):
    wb = Workbook()
    ws = wb.active

    # Row height
    ws.row_dimensions[1].height = 118

    for i in range(2, 6 + 1):
        ws.row_dimensions[i].height = 30

    ws.row_dimensions[7].height = 60
    ws.row_dimensions[3].height = 50

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 20.0
    for i in range(ord('C'), ord('C')+16):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)
    data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_r_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    # Img
    img = Image("myems.png")
    ws.add_image(img, 'B1')

    # Title
    ws['B3'].font = name_font
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = 'Name:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = name

    ws['D3'].font = name_font
    ws['D3'].alignment = b_r_alignment
    ws['D3'] = 'Period:'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = period_type

    ws['F3'].font = name_font
    ws['F3'].alignment = b_r_alignment
    ws['F3'] = 'Date:'
    ws['G3'].border = b_border
    ws['G3'].alignment = b_c_alignment
    ws['G3'].font = name_font
    ws['G3'] = reporting_start_datetime_local + "__" + reporting_end_datetime_local
    ws.merge_cells("G3:H3")

    ################################################
    # First: 趋势
    # 6: title
    # 7: table title
    # 8~ table_data
    ################################################
    has_data_flag = True
    report = result['reporting_period']
    if "names" not in report.keys() or report['names'] is None or len(report['names']) == 0:
        has_data_flag = False

    if "timestamps" not in report.keys() or report['timestamps'] is None or len(report['timestamps']) == 0:
        has_data_flag = False

    if "values" not in report.keys() or report['values'] is None or len(report['values']) == 0:
        has_data_flag = False
    ca = report['names']
    ca_len = len(ca)
    temp_max_row = 0
    times = report['timestamps']
    if has_data_flag:
        ws['B6'].font = title_font
        ws['B6'] = name + ' 趋势'

        ws['B7'].fill = table_fill
        ws['B7'].border = f_border
        ws['B7'].alignment = c_c_alignment
        ws['B7'] = '时间'
        time = times[0]
        has_data = False
        max_row = 0
        if len(time) > 0:
            has_data = True
            max_row = 8 + len(time)
            print("max_row", max_row)
            temp_max_row = max_row
        if has_data:
            for i in range(0, len(time)):
                col = 'B'
                row = str(8 + i)
                # col = chr(ord('B') + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = time[i]
                ws[col + row].border = f_border

            for i in range(0, ca_len):
                # 38 title
                col = chr(ord('C') + i)

                ws[col + '7'].fill = table_fill
                ws[col + '7'].font = title_font
                ws[col + '7'].alignment = c_c_alignment
                ws[col + '7'] = report['names'][i]
                ws[col + '7'].border = f_border

                # 39 data
                time = times[i]
                time_len = len(time)

                for j in range(0, time_len):
                    row = str(8 + j)
                    # col = chr(ord('B') + i)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = round(report['values'][i][j], 0)
                    ws[col + row].border = f_border
                # bar
                # 39~: bar
                bar = BarChart()
                labels = Reference(ws, min_col=2, min_row=8, max_row=max_row + 1)
                bar_data = Reference(ws, min_col=3 + i, min_row=7, max_row=max_row + 1)  # openpyxl bug
                bar.add_data(bar_data, titles_from_data=True)
                bar.set_categories(labels)
                bar.height = 5.25  # cm 1.05*5 1.05cm = 30 pt
                bar.width = 36
                # pie.title = "Pies sold by category"
                bar.dLbls = DataLabelList()
                # bar.dLbls.showCatName = True  # label show
                # bar.dLbls.showVal = True  # val show
                bar.dLbls.showPercent = True  # percent show
                # s1 = CharacterProperties(sz=1800)     # font size *100
                chart_col = chr(ord('B'))
                chart_cell = chart_col + str(max_row + 2 + 10*i)
                print("chart_cell", chart_cell)
                ws.add_chart(bar, chart_cell)
    else:
        pass

    for i in range(8, temp_max_row + 1 + 1):
        ws.row_dimensions[i].height = 20

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename